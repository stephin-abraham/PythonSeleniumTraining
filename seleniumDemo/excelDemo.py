import openpyxl

book = openpyxl.load_workbook("C:/Users/Stephin Philip A/Downloads/download.xlsx")
sheet = book.active
Dict = {}
cell = sheet.cell(row=1, column=2)
print(cell.value)
sheet.cell(row=1, column=2).value = "700"

print(sheet.cell(row=1, column=2).value)

print(sheet.max_row)

print(sheet.max_column)

print(sheet.min_column)

#print (sheet['A5'].value)
for i in range(1, sheet.max_row + 1):
    if sheet.cell(row = i, column= 1).value == 'Testcase2':

        for j in range(2, sheet.max_column + 1):
            Dict[sheet.cell(row = 1, column=j).value] = sheet.cell(row = i, column=j).value

print(Dict)